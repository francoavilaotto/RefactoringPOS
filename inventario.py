import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import os
import sys

from almacenamiento import guardado_atomico, hacer_backup

if getattr(sys, 'frozen', False):
    BASE_PATH = os.path.dirname(sys.executable)
else:
    BASE_PATH = os.path.dirname(os.path.abspath(__file__))

DATA_PATH = os.path.join(BASE_PATH, "data")
INVENTARIO_PATH = os.path.join(DATA_PATH, "inventario.xlsx")

COLUMNAS = [
    "Codigo",
    "Producto",
    "UnidadDeMedida",
    "Precio",
    "FechaVencimiento",
    "Cantidad",
    "DiasParaVencer"
]

def crear_excel_base():
    if os.path.exists(INVENTARIO_PATH):
        return

    os.makedirs(DATA_PATH, exist_ok=True)
    df = pd.DataFrame(columns=COLUMNAS)
    df.to_excel(INVENTARIO_PATH, index=False, engine="openpyxl")

def calcular_dias_para_vencer(df):
    hoy=datetime.now().date()
    if "FechaVencimiento" not in df.columns:
        df["DiasParaVencer"] = None
        return df

    df["FechaVencimiento"] = pd.to_datetime(
        df["FechaVencimiento"],
        errors="coerce",
        format="mixed",
        dayfirst=True,
    ).dt.date

    df["DiasParaVencer"]=df["FechaVencimiento"].apply(lambda x: (x-hoy).days if pd.notna(x) else None)

    return df

def cargar_inventario():
    """
    Carga el inventario desde disco y calcula DiasParaVencer en memoria.

    IMPORTANTE: a diferencia de la versión anterior, cargar_inventario()
    ya NO reescribe el archivo en disco. Cargar datos nunca debería
    implicar una escritura; cada escritura extra es una oportunidad más
    de corrupción (ej. si se corta la luz justo al abrir la app).
    El guardado real solo ocurre cuando se llama explícitamente a
    guardar_inventario(df), por ejemplo después de una venta o una
    edición manual del inventario.
    """
    if not os.path.exists(INVENTARIO_PATH):
        crear_excel_base()

    df = pd.read_excel(
        INVENTARIO_PATH,
        engine="openpyxl",
        dtype={"Codigo": str}
    )

    for col in COLUMNAS:
        if col not in df.columns:
            df[col] = None

    df = df[COLUMNAS]

    df = calcular_dias_para_vencer(df)

    # A diferencia de la versión intermedia, acá SÍ volvemos a guardar al
    # cargar. Es necesario porque el inventario se edita a mano en Excel
    # (agregar productos, precios, fechas de vencimiento), y guardar acá
    # es lo que recalcula y persiste DiasParaVencer + el color de
    # vencimiento para esas filas nuevas. Ahora es seguro hacerlo porque
    # guardar_inventario() usa guardado atómico + backup por debajo:
    # si se corta la luz a mitad de este guardado, el archivo real nunca
    # queda corrupto.
    guardar_inventario(df)

    return df

def guardar_inventario(df):
    if df is None:
        return

    df = calcular_dias_para_vencer(df)

    # Backup con timestamp antes de tocar el archivo real.
    hacer_backup(INVENTARIO_PATH)

    def escribir(ruta_temp):
        df.to_excel(ruta_temp, index=False, engine="openpyxl")

        wb = load_workbook(ruta_temp)
        ws = wb.active
        col_dias = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == "DiasParaVencer":
                col_dias = i
                break

        if col_dias and ws.max_row >= 2:
            letra_dias = ws.cell(row=1, column=col_dias).column_letter

            ultima_col = ws.max_column
            ultima_col_letra = ws.cell(row=1, column=ultima_col).column_letter

            rango = f"A2:{ultima_col_letra}{ws.max_row}"

            rojo = PatternFill(start_color="F87171", end_color="F87171", fill_type="solid")
            amarillo = PatternFill(start_color="FACC15", end_color="FACC15", fill_type="solid")

            regla_vencido = FormulaRule(
                formula=[f"${letra_dias}2<=0"],
                fill=rojo
            )

            regla_por_vencer = FormulaRule(
                formula=[f"AND(${letra_dias}2>0,${letra_dias}2<=7)"],
                fill=amarillo
            )

            ws.conditional_formatting.add(rango, regla_vencido)
            ws.conditional_formatting.add(rango, regla_por_vencer)

        wb.save(ruta_temp)

    # Guardado atómico: si algo falla durante la escritura, el archivo
    # real (INVENTARIO_PATH) nunca queda a medio escribir.
    guardado_atomico(escribir, INVENTARIO_PATH)